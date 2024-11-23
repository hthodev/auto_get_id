import os
import subprocess
import time
import keyboard
from pywinauto import Application
from pywinauto.keyboard import send_keys
import openpyxl
import pyperclip
import pyautogui
import random
from src.cucumber import log, mrh, pth, hju, kng, bru, htm
from global_config import base_dir, file_path, console_image_paths, seach_image_paths, close_image_paths

codes = dict(
    realgoats="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    tomarket_ai_bot="copy+9decodeURIComponent+9sessionStorage.SourceTarget+0.split+9'#tgWebAppData='+0[1].split+9'&tgWebAppVersion='+0[0]+0",
    tomarket="copy+9decodeURIComponent+9sessionStorage.SourceTarget+0.split+9'#tgWebAppData='+0[1].split+9'&tgWebAppVersion='+0[0]+0",
    money_dogs="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    bitget="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    drops_coin="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    freedogs="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    pumpad="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    neuton="copy+9decodeURIComponent+9sessionStorage['tapps/launchParams']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    catsgang="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0",
    bums="copy+9decodeURIComponent+9sessionStorage['telegram-apps/launch-params']+0.split+9'tgWebAppData='+0[1].split+9'&tgWebAppStartParam'+0[0]+0"
)

def load_accounts_from_file(file_path):
    """Loads accounts from a text file and returns them as a list."""
    with open(file_path, 'r') as file:
        # Read lines from the file and convert them to integers
        accounts = [int(line.strip()) for line in file if line.strip().isdigit()]
    return accounts

def load_bots_from_file(file_path):
    """Loads bots from a text file and returns them as a list."""
    with open(file_path, 'r') as file:
        # Read lines from the file and remove extra spaces
        bots = [line.strip() for line in file if line.strip()]
    return bots

def launch_telegram(account_num):
    account_folder = os.path.join(base_dir, str(account_num))
    try:
        exe_file = os.path.join(account_folder, f"{account_num}.exe")
        subprocess.Popen(exe_file)
        log(f"Launching Telegram for account {account_num} from {kng}{exe_file}")
    except:
        exe_file = os.path.join(account_folder, f"Telegram.exe")
        subprocess.Popen(exe_file)
        log(f"Launching Telegram for account {account_num} from {exe_file}...")

    time.sleep(15)  # Delay to wait for Telegram to launch
    log(hju + "Telegram launched.")

    return Application().connect(path=exe_file), exe_file

def close_telegram(account_num, exe_file):
    log(f"Closing Telegram for account {account_num}...")
    subprocess.call(f'taskkill /F /IM {os.path.basename(exe_file)}', shell=True)
    time.sleep(4)  # Delay to close

def write_to_excel(account_num, bot_name, row):
    # Open the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get all values from the first row
    first_row_values = [cell.value for cell in sheet[1]]

    # Check if bot_name exists in the first row
    if bot_name in first_row_values:
        # Find the column index for bot_name
        col_index = first_row_values.index(bot_name) + 1  # +1 for correct indexing in openpyxl
    else:
        # Add the bot name to the first row and get the index of the new column
        col_index = len(first_row_values) + 1  # New index for the next column
        sheet.cell(row=1, column=col_index, value=bot_name)

    # Copy data from clipboard
    time.sleep(1)  # Small delay before pasting
    tg_webapp_data = pyperclip.paste()

    # Add data to the table
    sheet["A1"] = "account_num"
    sheet[f'A{row}'] = account_num  # Account number
    sheet.cell(row=row, column=col_index, value=tg_webapp_data)  # Data from clipboard

    # Save the file
    workbook.save(file_path)
    log(hju + f"Data for account {account_num} successfully written to {file_path}")

def click_console(bot_name):
    # Search and click on the "Console" tab
    log("Searching for 'Console' tab on the screen...")

    if bot_name == 'agent301bot':
        time.sleep(5)
    else:
        time.sleep(4)

    for image_path in console_image_paths:
        try:
            console_location = pyautogui.locateOnScreen(image_path, confidence=0.8)
            if console_location:  # If the image is found, exit the loop
                log(hju + f"Console found at coordinates: {console_location}")
                break
        except Exception as e:
            log(f'Error while searching for image {image_path}: {e}')

    time.sleep(1)

    if console_location:
        log("Console tab found, clicking on it...")
        pyautogui.click(pyautogui.center(console_location))
        time.sleep(1)
    else:
        log("Failed to find Console tab.")
        return

def kucoin():
    kucoin = r"pic\support\kucoin_button.png"
    time.sleep(9)

    try:
        for _ in range(4):
            kucoin_click = pyautogui.locateOnScreen(kucoin, confidence=0.8)
            time.sleep(1)
            if kucoin_click:
                log("Searching for KuCoin button")
                pyautogui.click(pyautogui.center(kucoin_click), duration=0.5)
                time.sleep(1)
            else:
                log("Failed to find KuCoin button")

        time.sleep(3)

    except Exception as e:
        log(f'Error during KuCoin execution {e}')
        return


def etherdrops():
    kucoin = r"pic\support\ether_b1.png"
    ether2 = r"pic\support\ether_b2.png"
    time.sleep(3)

    try:
        for _ in range(3):
            kucoin_click = pyautogui.locateOnScreen(kucoin, confidence=0.8)
            time.sleep(1)
            if kucoin_click:
                log("Searching for ether button")
                pyautogui.click(pyautogui.center(kucoin_click), duration=0.5)
                time.sleep(1)
            else:
                log("Failed to find ether button")
        time.sleep(1)
    except Exception as e:
        log(f'Error during ether execution {e}')

    try:
        kucoin_click2 = pyautogui.locateOnScreen(ether2, confidence=0.8)
        time.sleep(1)
        if kucoin_click2:
            log("Searching for ether2 button")
            pyautogui.click(pyautogui.center(kucoin_click2), duration=0.5)
            time.sleep(2)
        else:
            log("Failed to find ether2 button")
    except Exception as e:
        log(f'Error during ether2 execution {e}')
        return


def fastmint():
    bird_button1 = r"pic\support\fastmint_b1.png"
    bird_button2 = r"pic\support\fastmint_b2.png"
    bird_button3 = r"pic\support\fastmint_b3.png"
    time.sleep(6)

    try:
        pyautogui.scroll(-300)
        time.sleep(1)
        log("Searching for fastmint_b1")
        duck_click = pyautogui.locateOnScreen(bird_button1, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking fastmint_b1")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to find fastmint_b1")
        time.sleep(2)

        pyautogui.scroll(-300)
        time.sleep(1)
        log("Searching for fastmint_b2")
        duck_click = pyautogui.locateOnScreen(bird_button2, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking fastmint_b2")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to fastmint_b2")
        time.sleep(2)

        pyautogui.scroll(-300)
        time.sleep(1)
        log("Searching for fastmint_b3")
        duck_click = pyautogui.locateOnScreen(bird_button3, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking fastmint_b3")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to fastmint_b3")
        time.sleep(2)

    except Exception as e:
        log(f'Error during fastmint execution {e}')
        return


def birds():
    bird_button1 = r"pic\support\bird_button1.png"
    bird_button2 = r"pic\support\bird_button2.png"
    time.sleep(2)

    try:
        log("Searching for bird button")
        duck_click = pyautogui.locateOnScreen(bird_button1, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking bird button")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to find bird button")
        time.sleep(5)

        log("Searching for bird button 2")
        duck_click = pyautogui.locateOnScreen(bird_button2, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking bird button 2")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to find bird button 2")
        time.sleep(3)

    except Exception as e:
        log(f'Error during birds execution {e}')
        return

def kiloextrade():
    bird_button1 = r"pic\support\kilo_b1.png"

    try:
        log("Searching for kiloextrade button")
        duck_click = pyautogui.locateOnScreen(bird_button1, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking kiloextrade button")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to find kiloextrade button")
        time.sleep(2)

    except Exception as e:
        log(f'Error during kiloextrade execution {e}')
        return

def bitget():
    bitget_button1 = r"pic\support\bitget1.png"
    bitget_button2 = r"pic\support\bitget2.png"
    time.sleep(2)
    biget_code = ""

    try:
        log("Searching for bird button")
        duck_click = pyautogui.locateOnScreen(bitget_button1, confidence=0.8)
        time.sleep(1)
        if duck_click:
            log("Clicking bird button")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to find bird button")
        time.sleep(1)

        log("Searching for bird button 2")
        duck_click = pyautogui.locateOnScreen(bitget_button2, confidence=0.9)
        time.sleep(1)
        if duck_click:
            log("Clicking bird button 2")
            pyautogui.click(pyautogui.center(duck_click), duration=0.5)
        else:
            log("Failed to find bird button 2")
        time.sleep(3)

        send_keys(biget_code)
        time.sleep(3)

        send_keys(biget_code)
        time.sleep(5)
    except Exception as e:
        log(f'Error during birds execution {e}')
        return

def duck_chain():
    duck = r"pic\support\duck_button.png"
    time.sleep(3)

    try:
        try:
            log("Searching for Duck button")
            duck_click = pyautogui.locateOnScreen(duck, confidence=0.8)
            time.sleep(1)
            if duck_click:
                log("Clicking Duck button")
                pyautogui.click(pyautogui.center(duck_click), duration=0.5)
            else:
                log("Failed to find Duck button")

        except Exception as e:
            log(f'Error while clicking Duck button {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Error during Duck execution {e}')
        return

def coub():
    duck = r"pic\support\coub_button.png"
    time.sleep(1)

    try:
        try:
            log("Searching for Duck button")
            duck_click = pyautogui.locateOnScreen(duck, confidence=0.8)
            time.sleep(1)
            if duck_click:
                log("Clicking Duck button")
                pyautogui.click(pyautogui.center(duck_click), duration=0.5)
            else:
                log("Failed to find Duck button")

        except Exception as e:
            log(f'Error while clicking Duck button {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Error during Duck execution {e}')
        return

def money_dogs():
    money = r"pic\support\money.png"
    time.sleep(3)

    try:
        try:
            log("Searching for Duck button")
            money_click = pyautogui.locateOnScreen(money, confidence=0.8)
            time.sleep(1)
            if money_click:
                log("Clicking Duck button")
                pyautogui.click(pyautogui.center(money_click), duration=0.5)
            else:
                log("Failed to find Duck button")

        except Exception as e:
            log(f'Error while clicking Duck button {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Error during Duck execution {e}')
        return

def padton():
    money = r"pic\support\padton_button.png"
    time.sleep(3)

    try:
        try:
            log("Searching for padton_button")
            money_click = pyautogui.locateOnScreen(money, confidence=0.8)
            time.sleep(1)
            if money_click:
                log("Clicking padton_button")
                pyautogui.click(pyautogui.center(money_click), duration=0.5)
            else:
                log("Failed to find padton_button")

        except Exception as e:
            log(f'Error while clicking padton_button {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Error during padton_button {e}')
        return


def pumpad():
    fabrika1 = r"pic\support\pumpad_b1.png"
    fabrika3 = r"pic\support\pumpad_b3.png"
    bot_image_paths = [r"pic\support\pumpad_b2.png",
                       r"pic\support\pumpad_b4.png"]
    time.sleep(1)


    try:
        log("Searching for pumpad button")
        fabrika_click = pyautogui.locateOnScreen(fabrika1, confidence=0.8)
        time.sleep(1)
        log("Clicking pumpad button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(3)

        for image_path in bot_image_paths:
            try:
                link_location = pyautogui.locateOnScreen(image_path, confidence=0.8)
                if link_location:  # If the image is found, exit the loop
                    break
            except Exception as e:
                log(f'Error while searching for image {image_path}: {e}')

        log("Clicking pumpad button")
        pyautogui.click(pyautogui.center(link_location), duration=0.5)
        time.sleep(2)

        keyboard.press_and_release('alt+tab')

        time.sleep(1)
        log("Searching for pumpad button")
        fabrika_click = pyautogui.locateOnScreen(fabrika3, confidence=0.8)
        time.sleep(1)
        log("Clicking pumpad button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(3)



    except Exception as e:
        log(f'Error during Duck execution {e}')
        return


def fabrika():
    fabrika1 = r"pic\support\fabrika_b1.png"
    fabrika2 = r"pic\support\fabrika_b2.png"
    fabrika3 = r"pic\support\fabrika_b3.png"
    fabrika4 = r"pic\support\fabrika_b4.png"
    time.sleep(1)


    try:
        log("Searching for fabrika button")
        fabrika_click = pyautogui.locateOnScreen(fabrika1, confidence=0.8)
        time.sleep(1)
        log("Clicking fabrika button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(2)

        log("Searching for fabrika button")
        fabrika_click = pyautogui.locateOnScreen(fabrika2, confidence=0.8)
        time.sleep(1)
        log("Clicking fabrika button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(2)

        log("Searching for fabrika button")
        fabrika_click = pyautogui.locateOnScreen(fabrika3, confidence=0.8)
        time.sleep(1)
        log("Clicking fabrika button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(2)

        log("Searching for fabrika button")
        fabrika_click = pyautogui.locateOnScreen(fabrika3, confidence=0.8)
        time.sleep(1)
        log("Clicking fabrika button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(2)

        log("Searching for fabrika button")
        fabrika_click = pyautogui.locateOnScreen(fabrika4, confidence=0.8)
        time.sleep(1)
        log("Clicking fabrika button")
        pyautogui.click(pyautogui.center(fabrika_click), duration=0.5)
        time.sleep(2)


    except Exception as e:
        log(f'Error during Duck execution {e}')
        return

def catsdogs():
    catsdogs = r"pic\support\catsdogs.png"
    catsdogs2 = r"pic\support\catsdogs2.png"
    catsdogs3 = r"pic\support\catsdogs3.png"
    time.sleep(1)

    try:
        try:
            log("Searching for catsdogs button")
            catsdogs_click = pyautogui.locateOnScreen(catsdogs, confidence=0.8)
            time.sleep(1)
            log("Clicking catsdogs button")
            pyautogui.click(pyautogui.center(catsdogs_click), duration=0.5)
            time.sleep(2)

            log("Searching for catsdogs button2")
            if random.random() < 0.5:
                catsdogs_click = pyautogui.locateOnScreen(catsdogs2, confidence=0.8)
            else:
                catsdogs_click = pyautogui.locateOnScreen(catsdogs3, confidence=0.8)
            time.sleep(1)
            log("Clicking catsdogs button2")
            pyautogui.click(pyautogui.center(catsdogs_click), duration=0.5)
            time.sleep(2)

            log("Searching for catsdogs button3")
            catsdogs_click = pyautogui.locateOnScreen(catsdogs, confidence=0.8)
            time.sleep(1)
            log("Clicking catsdogs button3")
            pyautogui.click(pyautogui.center(catsdogs_click), duration=0.5)
            time.sleep(2)

        except Exception as e:
            log(f'Error while clicking Duck button {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Error during Duck execution {e}')
        return

def match_quest():
    match1 = r"pic\support\match_b1.png"
    match2 = r"pic\support\match_b2.png"

    time.sleep(1)

    try:
        try:
            match_click1 = pyautogui.locateOnScreen(match1, confidence=0.7)
            time.sleep(1)
            if match_click1:
                log("Searching for Match1 button")
                pyautogui.click(pyautogui.center(match_click1))
            else:
                log("Failed to find Match1 button")

        except Exception as e:
            log(f'Error while clicking Match1 button {e}')
            return

        time.sleep(5)

        try:
            match_click2 = pyautogui.locateOnScreen(match2, confidence=0.7)
            time.sleep(1)
            if match_click2:
                log("Searching for Match2 button")
                pyautogui.click(pyautogui.center(match_click2))
            else:
                log("Failed to find Match2 button")

        except Exception as e:
            log(f'Error while clicking Match2 image {e}')
            return

        time.sleep(5)

    except Exception as e:
        log(f'Error during Match execution {e}')
        return

def open_bot(app):
    log("Focusing on the Telegram window...")
    main_window = app.top_window()
    main_window.set_focus()

    time.sleep(1.5)

    log("Opening search...")
    main_window.type_keys('^f')  # Ctrl + F

    time.sleep(2.5)

    log(f"Entering group name: ...")

    main_window.type_keys('ref_airdrop')

    time.sleep(2.5)

    log("Pressing down key to select the bot and Enter...")

    main_window.type_keys('{DOWN}')

    time.sleep(1.5)

    main_window.type_keys('{ENTER}')

    time.sleep(1.5)

def close_app():
    try:
        for image_path in close_image_paths:
            try:
                close_location = pyautogui.locateOnScreen(image_path, confidence=0.7)
                if close_location:
                    time.sleep(1)
                    pyautogui.click(pyautogui.center(close_location))
                    log(f'Clicking application close button')
                    time.sleep(1)
                    break
            except:
                log(f'')

    except:
        log('')

def click_seach():
    time.sleep(1)  # Give time for the link to appear

    for image_path in seach_image_paths:
        try:
            seach_location = pyautogui.locateOnScreen(image_path, confidence=0.8)
            time.sleep(1)
            if seach_location:  # If the image is found, exit the loop
                break
        except Exception as e:
            log(f'Error while searching for image {image_path}: {e}')

    pyautogui.click(pyautogui.center(seach_location))
    time.sleep(1)  # Delay before pressing Enter